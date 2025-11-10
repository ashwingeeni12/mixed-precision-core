"""
Create comprehensive Excel spreadsheet from test results.
Generates multiple sheets with formatted data, summaries, and analysis.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import argparse

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import ScatterChart, Reference, Series, BarChart
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not available. Will create basic Excel file without formatting.")

def create_summary_stats(df):
    """Create summary statistics grouped by precision and condition category."""
    summary_data = []

    for prec in ['int8', 'fp16', 'fp32']:
        for cat in ['low', 'medium', 'high']:
            subset = df[(df['precision'] == prec) & (df['category'] == cat)]

            if len(subset) == 0:
                continue

            summary_data.append({
                'precision': prec,
                'category': cat,
                'num_cases': len(subset),
                'mean_mae': subset['mae'].mean(),
                'std_mae': subset['mae'].std(),
                'mean_rmse': subset['rmse'].mean(),
                'std_rmse': subset['rmse'].std(),
                'mean_rel_rmse': subset['rel_rmse'].mean(),
                'std_rel_rmse': subset['rel_rmse'].std(),
                'mean_max_error': subset['max_abs_error'].mean(),
                'max_max_error': subset['max_abs_error'].max(),
                'mean_snr_db': subset['snr_db'].mean(),
                'min_snr_db': subset['snr_db'].min(),
                'mean_correlation': subset['correlation'].mean(),
                'min_correlation': subset['correlation'].min(),
                'mean_acc_1pct': subset['acc_1pct'].mean(),
                'mean_acc_5pct': subset['acc_5pct'].mean(),
                'mean_acc_10pct': subset['acc_10pct'].mean(),
                'mean_sim_time': subset['sim_time_sec'].mean(),
                'total_sim_time': subset['sim_time_sec'].sum(),
            })

    return pd.DataFrame(summary_data)

def create_precision_comparison(df):
    """Create precision comparison table."""
    comp_data = []

    for prec in ['int8', 'fp16', 'fp32']:
        subset = df[df['precision'] == prec]

        if len(subset) == 0:
            continue

        comp_data.append({
            'precision': prec,
            'total_cases': len(subset),
            'overall_mean_mae': subset['mae'].mean(),
            'overall_mean_rmse': subset['rmse'].mean(),
            'overall_mean_rel_rmse': subset['rel_rmse'].mean(),
            'overall_mean_snr_db': subset['snr_db'].mean(),
            'overall_mean_correlation': subset['correlation'].mean(),
            'overall_mean_acc_1pct': subset['acc_1pct'].mean(),
            'overall_mean_acc_5pct': subset['acc_5pct'].mean(),
            'best_snr_db': subset['snr_db'].max(),
            'worst_snr_db': subset['snr_db'].min(),
            'total_time_sec': subset['sim_time_sec'].sum(),
            'avg_time_per_case': subset['sim_time_sec'].mean(),
        })

    return pd.DataFrame(comp_data)

def create_condition_analysis(df):
    """Analyze impact of condition number on accuracy."""
    analysis_data = []

    for prec in ['int8', 'fp16', 'fp32']:
        prec_data = df[df['precision'] == prec].copy()

        if len(prec_data) == 0:
            continue

        # Compute correlation between condition number and error metrics
        try:
            corr_cond_mae = prec_data['actual_cond_A'].corr(prec_data['mae'])
            corr_cond_rmse = prec_data['actual_cond_A'].corr(prec_data['rmse'])
            corr_cond_rel_rmse = prec_data['actual_cond_A'].corr(prec_data['rel_rmse'])
        except:
            corr_cond_mae = np.nan
            corr_cond_rmse = np.nan
            corr_cond_rel_rmse = np.nan

        analysis_data.append({
            'precision': prec,
            'corr_cond_vs_mae': corr_cond_mae,
            'corr_cond_vs_rmse': corr_cond_rmse,
            'corr_cond_vs_rel_rmse': corr_cond_rel_rmse,
            'low_cond_mean_rmse': prec_data[prec_data['category'] == 'low']['rmse'].mean(),
            'med_cond_mean_rmse': prec_data[prec_data['category'] == 'medium']['rmse'].mean(),
            'high_cond_mean_rmse': prec_data[prec_data['category'] == 'high']['rmse'].mean(),
            'low_cond_mean_snr': prec_data[prec_data['category'] == 'low']['snr_db'].mean(),
            'med_cond_mean_snr': prec_data[prec_data['category'] == 'medium']['snr_db'].mean(),
            'high_cond_mean_snr': prec_data[prec_data['category'] == 'high']['snr_db'].mean(),
        })

    return pd.DataFrame(analysis_data)

def format_worksheet(ws, df, title=""):
    """Apply formatting to worksheet if openpyxl is available."""
    if not HAS_OPENPYXL:
        return

    # Header formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', type=str, default='../results/comprehensive_results.csv',
                        help='Input CSV file with test results')
    parser.add_argument('--output', type=str, default='../results/comprehensive_analysis.xlsx',
                        help='Output Excel file')
    args = parser.parse_args()

    input_file = Path(args.input)
    output_file = Path(args.output)

    if not input_file.exists():
        print(f"Error: Input file {input_file} not found!")
        return

    print(f"Reading results from {input_file}...")
    df = pd.read_csv(input_file)

    print(f"Loaded {len(df)} test results")

    # Filter out failed runs
    df_success = df[df['status'] == 'success'].copy()
    df_failed = df[df['status'] != 'success'].copy()

    print(f"  Success: {len(df_success)}")
    print(f"  Failed: {len(df_failed)}")

    # Create Excel writer
    print(f"\nCreating Excel file: {output_file}...")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: All raw data
        print("  Creating sheet: Raw Data")
        df.to_excel(writer, sheet_name='Raw Data', index=False)

        # Sheet 2: Successful runs only
        print("  Creating sheet: Successful Runs")
        df_success.to_excel(writer, sheet_name='Successful Runs', index=False)

        # Sheet 3: Summary statistics
        print("  Creating sheet: Summary Statistics")
        summary = create_summary_stats(df_success)
        summary.to_excel(writer, sheet_name='Summary Statistics', index=False)

        # Sheet 4: Precision comparison
        print("  Creating sheet: Precision Comparison")
        precision_comp = create_precision_comparison(df_success)
        precision_comp.to_excel(writer, sheet_name='Precision Comparison', index=False)

        # Sheet 5: Condition number analysis
        print("  Creating sheet: Condition Analysis")
        cond_analysis = create_condition_analysis(df_success)
        cond_analysis.to_excel(writer, sheet_name='Condition Analysis', index=False)

        # Sheet 6: Error distribution by precision
        print("  Creating sheet: Error Distribution")
        error_dist = df_success.groupby('precision').agg({
            'mae': ['min', 'max', 'mean', 'median', 'std'],
            'rmse': ['min', 'max', 'mean', 'median', 'std'],
            'rel_rmse': ['min', 'max', 'mean', 'median', 'std'],
            'max_abs_error': ['min', 'max', 'mean', 'median', 'std'],
        }).round(6)
        error_dist.to_excel(writer, sheet_name='Error Distribution')

        # Sheet 7: Best and worst cases
        print("  Creating sheet: Best and Worst Cases")
        best_worst_data = []

        for prec in ['int8', 'fp16', 'fp32']:
            subset = df_success[df_success['precision'] == prec]
            if len(subset) == 0:
                continue

            best_idx = subset['rmse'].idxmin()
            worst_idx = subset['rmse'].idxmax()

            best_worst_data.append({
                'type': 'BEST',
                'precision': prec,
                'case_id': subset.loc[best_idx, 'case_id'],
                'category': subset.loc[best_idx, 'category'],
                'mae': subset.loc[best_idx, 'mae'],
                'rmse': subset.loc[best_idx, 'rmse'],
                'rel_rmse': subset.loc[best_idx, 'rel_rmse'],
                'snr_db': subset.loc[best_idx, 'snr_db'],
                'correlation': subset.loc[best_idx, 'correlation'],
            })

            best_worst_data.append({
                'type': 'WORST',
                'precision': prec,
                'case_id': subset.loc[worst_idx, 'case_id'],
                'category': subset.loc[worst_idx, 'category'],
                'mae': subset.loc[worst_idx, 'mae'],
                'rmse': subset.loc[worst_idx, 'rmse'],
                'rel_rmse': subset.loc[worst_idx, 'rel_rmse'],
                'snr_db': subset.loc[worst_idx, 'snr_db'],
                'correlation': subset.loc[worst_idx, 'correlation'],
            })

        best_worst_df = pd.DataFrame(best_worst_data)
        best_worst_df.to_excel(writer, sheet_name='Best and Worst Cases', index=False)

        # Sheet 8: Failed runs (if any)
        if len(df_failed) > 0:
            print("  Creating sheet: Failed Runs")
            df_failed.to_excel(writer, sheet_name='Failed Runs', index=False)

    # Apply formatting if openpyxl is available
    if HAS_OPENPYXL:
        print("\nApplying formatting...")
        wb = load_workbook(output_file)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            format_worksheet(ws, None)

        wb.save(output_file)

    print(f"\n{'='*80}")
    print(f"Excel report created successfully!")
    print(f"Output file: {output_file}")
    print(f"{'='*80}")

    # Print summary
    print("\n=== SUMMARY ===")
    print(f"Total test cases: {len(df)}")
    print(f"Successful: {len(df_success)}")
    print(f"Failed: {len(df_failed)}")

    if len(df_success) > 0:
        print("\nOverall Statistics:")
        for prec in ['int8', 'fp16', 'fp32']:
            subset = df_success[df_success['precision'] == prec]
            if len(subset) > 0:
                print(f"\n{prec.upper()}:")
                print(f"  Cases: {len(subset)}")
                print(f"  Mean RMSE: {subset['rmse'].mean():.6f}")
                print(f"  Mean Rel RMSE: {subset['rel_rmse'].mean():.6f}")
                print(f"  Mean SNR: {subset['snr_db'].mean():.2f} dB")
                print(f"  Mean Correlation: {subset['correlation'].mean():.6f}")

if __name__ == '__main__':
    main()
