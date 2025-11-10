"""
Create comprehensive Excel inventory of all test cases.
"""

import pandas as pd
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

ROOT = Path(__file__).resolve().parents[1]
TEST_CASES_DIR = ROOT / "test_cases"
RESULTS_DIR = ROOT / "results"
RESULTS_DIR.mkdir(exist_ok=True)

def create_test_inventory():
    """Create Excel file with all test case information."""

    # Load all metadata
    metadata_file = TEST_CASES_DIR / "all_test_metadata.json"
    with open(metadata_file, 'r') as f:
        all_metadata = json.load(f)

    print(f"Loaded {len(all_metadata)} test cases")

    # Create DataFrame
    df = pd.DataFrame(all_metadata)

    # Reorder columns for better presentation
    columns_order = [
        'case_id', 'category', 'precision',
        'requested_cond_A', 'requested_cond_B',
        'actual_cond_A', 'actual_cond_B',
        'M', 'K', 'N',
        'A_min', 'A_max', 'A_mean', 'A_std',
        'B_min', 'B_max', 'B_mean', 'B_std',
        'C_ref_min', 'C_ref_max', 'C_ref_mean', 'C_ref_std',
        'C_ref_norm', 'seed'
    ]

    df = df[columns_order]

    # Create Excel file
    output_file = RESULTS_DIR / "test_inventory.xlsx"

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: All test cases
        df.to_excel(writer, sheet_name='All Test Cases', index=False)

        # Sheet 2: Summary by category
        summary_data = []
        for prec in ['int8', 'fp16', 'fp32']:
            for cat in ['low', 'medium', 'high']:
                subset = df[(df['precision'] == prec) & (df['category'] == cat)]
                if len(subset) > 0:
                    summary_data.append({
                        'precision': prec,
                        'category': cat,
                        'num_cases': len(subset),
                        'avg_cond_A': subset['actual_cond_A'].mean(),
                        'min_cond_A': subset['actual_cond_A'].min(),
                        'max_cond_A': subset['actual_cond_A'].max(),
                        'avg_cond_B': subset['actual_cond_B'].mean(),
                        'min_cond_B': subset['actual_cond_B'].min(),
                        'max_cond_B': subset['actual_cond_B'].max(),
                    })

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Sheet 3: Template for results
        results_columns = [
            'case_id', 'category', 'precision',
            'M', 'K', 'N',
            'actual_cond_A', 'actual_cond_B',
            # Placeholders for simulation results
            'MAE', 'RMSE', 'Max_Abs_Error',
            'Rel_MAE', 'Rel_RMSE', 'Max_Rel_Error',
            'Norm_Rel_Error',
            'P50_Error', 'P90_Error', 'P95_Error', 'P99_Error',
            'SNR_dB', 'Correlation',
            'Acc_1pct', 'Acc_5pct', 'Acc_10pct',
            'Sim_Time_sec', 'Status', 'Notes'
        ]

        results_df = df[['case_id', 'category', 'precision', 'M', 'K', 'N', 'actual_cond_A', 'actual_cond_B']].copy()

        # Add empty columns for results
        for col in results_columns[8:]:
            results_df[col] = ''

        results_df.to_excel(writer, sheet_name='Results Template', index=False)

        # Sheet 4: By precision
        for prec in ['int8', 'fp16', 'fp32']:
            prec_df = df[df['precision'] == prec].copy()
            prec_df.to_excel(writer, sheet_name=f'{prec.upper()}', index=False)

    # Format workbook
    if HAS_OPENPYXL:
        wb = load_workbook(output_file)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row
            ws.freeze_panes = 'A2'

        wb.save(output_file)

    print(f"\nCreated: {output_file}")
    print(f"Sheets: {len(all_metadata)} test cases across multiple sheets")

    return output_file

if __name__ == '__main__':
    create_test_inventory()
